"""Exportação XLSX do resultado de conciliação."""

from __future__ import annotations

import duckdb
import openpyxl

from altool.engine.conciliacao import KeyDef, conciliar_multichave
from altool.engine.export import MONEY_FMT, export_resultado_xlsx


def _con_resultado():  # type: ignore[no-untyped-def]
    con = duckdb.connect()
    con.execute("CREATE TABLE base_a(chave VARCHAR, valor VARCHAR)")
    con.execute("CREATE TABLE base_b(chave VARCHAR, valor VARCHAR)")
    con.executemany("INSERT INTO base_a VALUES (?,?)", [("K1", "1000"), ("K2", "500")])
    con.executemany("INSERT INTO base_b VALUES (?,?)", [("K1", "1000"), ("K3", "300")])
    conciliar_multichave(
        con, [KeyDef("CHAVE_1", ["chave"], ["chave"])],
        value_col_a="valor", value_col_b="valor",
    )
    return con


def test_export_gera_xlsx_valido(tmp_path) -> None:  # type: ignore[no-untyped-def]
    con = _con_resultado()
    out = tmp_path / "resultado.xlsx"
    n = export_resultado_xlsx(con, "conciliacao_result", str(out))
    assert out.exists()
    # K1 (A+B), K2 (só A → 03), K3 (só B → 03) = 4 linhas nível-linha.
    assert n == 4

    wb = openpyxl.load_workbook(out)
    ws = wb["resultado"]
    headers = [c.value for c in ws[1]]
    assert headers[:4] == ["origem", "row_id", "key_id", "status"]
    assert ws.max_row == 5  # 1 header + 4 dados


def test_export_formatacao_monetaria(tmp_path) -> None:  # type: ignore[no-untyped-def]
    con = _con_resultado()
    out = tmp_path / "r.xlsx"
    export_resultado_xlsx(con, "conciliacao_result", str(out))
    wb = openpyxl.load_workbook(out)
    ws = wb["resultado"]
    headers = [c.value for c in ws[1]]
    vcol = headers.index("value_a") + 1  # 1-based

    money_cell = ws.cell(row=2, column=vcol)
    assert money_cell.number_format == MONEY_FMT
    assert isinstance(money_cell.value, (int, float))  # escrito como número, não texto
