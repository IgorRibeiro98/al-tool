#!/usr/bin/env python3
"""Streamed spreadsheet converter: .xlsb <-> .xlsx and JSONL export.

Improvements in this refactor:
- Use pathlib for path handling and ensure output directories exist before writing.
- Add logging and proper exit codes on failure.
- Add type hints and clearer JSON normalization for numeric/date types.
- PERFORMANCE: Buffered I/O for faster writes, batch processing for memory efficiency.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import io
import gc
from decimal import Decimal
import datetime as _dt
from pathlib import Path
from typing import Any, Iterator, List

from pyxlsb import open_workbook
from openpyxl import Workbook
from openpyxl import load_workbook


logger = logging.getLogger("xlsb_to_xlsx")
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

# ============================================================================
# PERFORMANCE CONSTANTS - Optimized for 8GB RAM Windows machines
# ============================================================================
# Buffer size for file I/O (larger = fewer syscalls = faster)
WRITE_BUFFER_SIZE = 8 * 1024 * 1024  # 8MB buffer
# Rows to accumulate before flushing to disk
FLUSH_INTERVAL = 10000  # Flush every 10k rows
# Progress logging interval
PROGRESS_LOG_INTERVAL = 50000  # Log every 50k rows


def _normalize_value_for_json(v: Any) -> Any:
    """Normalize cell values to JSON-serializable primitives.

    - numbers -> {"__num__": "<decimal-string>"}
    - Decimal -> {"__num__": "<decimal-string>"}
    - dates/datetimes -> ISO string
    - bool -> bool
    - None -> None
    - otherwise: return as-is if json-serializable, else str(v)
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        # Convert floats/ints to Decimal via string to avoid float artifacts
        d = Decimal(str(v))
        return {"__num__": format(d, "f")}
    if isinstance(v, Decimal):
        return {"__num__": format(v, "f")}
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    try:
        json.dumps(v)
        return v
    except TypeError:
        return str(v)


def _ensure_parent_dir(path: Path) -> None:
    parent = path.parent
    if not parent.exists():
        parent.mkdir(parents=True, exist_ok=True)


def convert_xlsb_to_xlsx(inpath: str, outpath: str, sheet_index: int = 1) -> None:
    in_path = Path(inpath)
    out_path = Path(outpath)
    _ensure_parent_dir(out_path)

    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title=f"Sheet{sheet_index}")

    with open_workbook(str(in_path)) as wb:
        sheet_names = list(wb.sheets)
        if sheet_index < 1 or sheet_index > len(sheet_names):
            raise ValueError(f"sheet_index {sheet_index} out of range (1..{len(sheet_names)})")
        sheet_name = sheet_names[sheet_index - 1]
        with wb.get_sheet(sheet_name) as sheet:
            for row in sheet.rows():
                out_row = []
                for cell in row:
                    if cell is None or getattr(cell, "v", None) is None:
                        out_row.append(None)
                    else:
                        out_row.append(cell.v)
                ws_out.append(out_row)

    wb_out.save(str(out_path))


def convert_xlsb_to_jsonl(inpath: str, outpath: str, sheet_index: int = 1) -> None:
    """Convert XLSB to JSONL with buffered I/O for better performance."""
    in_path = Path(inpath)
    out_path = Path(outpath)
    _ensure_parent_dir(out_path)

    rows_written = 0
    # Use buffered writer for much faster I/O
    with open(str(out_path), "w", encoding="utf-8", buffering=WRITE_BUFFER_SIZE) as fout:
        with open_workbook(str(in_path)) as wb:
            sheet_names = list(wb.sheets)
            if sheet_index < 1 or sheet_index > len(sheet_names):
                raise ValueError(f"sheet_index {sheet_index} out of range (1..{len(sheet_names)})")
            sheet_name = sheet_names[sheet_index - 1]
            logger.info(f"Converting sheet '{sheet_name}' from {in_path.name}")
            
            with wb.get_sheet(sheet_name) as sheet:
                for row in sheet.rows():
                    out_row = []
                    for cell in row:
                        if cell is None or getattr(cell, "v", None) is None:
                            out_row.append(None)
                            continue
                        out_row.append(_normalize_value_for_json(cell.v))
                    fout.write(json.dumps(out_row, ensure_ascii=False) + "\n")
                    rows_written += 1
                    
                    # Progress logging for large files
                    if rows_written % PROGRESS_LOG_INTERVAL == 0:
                        logger.info(f"Processed {rows_written:,} rows...")
                    
                    # Periodic flush to prevent memory buildup
                    if rows_written % FLUSH_INTERVAL == 0:
                        fout.flush()
    
    # Force garbage collection after large conversion
    gc.collect()
    logger.info(f"Conversion complete: {rows_written:,} rows written to {out_path.name}")


def convert_xlsx_to_jsonl(inpath: str, outpath: str, sheet_index: int = 1) -> None:
    """Convert XLSX to JSONL with buffered I/O and optimized memory usage."""
    in_path = Path(inpath)
    out_path = Path(outpath)
    _ensure_parent_dir(out_path)

    rows_written = 0
    # Use buffered writer for much faster I/O
    with open(str(out_path), "w", encoding="utf-8", buffering=WRITE_BUFFER_SIZE) as fout:
        # read_only=True and data_only=True are critical for memory efficiency
        wb = load_workbook(filename=str(in_path), read_only=True, data_only=True)
        sheets = wb.sheetnames
        if sheet_index < 1 or sheet_index > len(sheets):
            raise ValueError(f"sheet_index {sheet_index} out of range (1..{len(sheets)})")
        sheet = wb[sheets[sheet_index - 1]]
        logger.info(f"Converting sheet '{sheets[sheet_index - 1]}' from {in_path.name}")
        
        for row in sheet.iter_rows(values_only=True):
            out_row = []
            for v in row:
                out_row.append(_normalize_value_for_json(v))
            fout.write(json.dumps(out_row, ensure_ascii=False) + "\n")
            rows_written += 1
            
            # Progress logging for large files
            if rows_written % PROGRESS_LOG_INTERVAL == 0:
                logger.info(f"Processed {rows_written:,} rows...")
            
            # Periodic flush to prevent memory buildup
            if rows_written % FLUSH_INTERVAL == 0:
                fout.flush()
        
        # Close workbook to release memory
        wb.close()
    
    # Force garbage collection after large conversion
    gc.collect()
    logger.info(f"Conversion complete: {rows_written:,} rows written to {out_path.name}")


def main(argv: list[str] | None = None) -> int:
    argv = list(argv) if argv is not None else sys.argv
    p = argparse.ArgumentParser(description="Convert .xlsb to .xlsx or JSONL (streaming)")
    p.add_argument("infile")
    p.add_argument("outfile")
    p.add_argument("--sheet", type=int, default=1, help="1-based sheet index to convert")
    p.add_argument("--jsonl", action="store_true", help="emit JSONL instead of XLSX (preserves numeric as Decimal strings)")
    args = p.parse_args(argv[1:])

    infile_lower = args.infile.lower()
    try:
        if args.jsonl:
            if infile_lower.endswith(".xlsb"):
                convert_xlsb_to_jsonl(args.infile, args.outfile, sheet_index=args.sheet)
            elif infile_lower.endswith(".xlsx") or infile_lower.endswith(".xls"):
                convert_xlsx_to_jsonl(args.infile, args.outfile, sheet_index=args.sheet)
            else:
                # fallback: try xlsb reader first, else attempt openpyxl
                try:
                    convert_xlsb_to_jsonl(args.infile, args.outfile, sheet_index=args.sheet)
                except Exception:
                    convert_xlsx_to_jsonl(args.infile, args.outfile, sheet_index=args.sheet)
        else:
            # default behavior: only support xlsb -> xlsx
            convert_xlsb_to_xlsx(args.infile, args.outfile, sheet_index=args.sheet)
    except Exception as exc:  # pragma: no cover - top-level safety
        logger.exception("Conversion failed: %s", exc)
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
