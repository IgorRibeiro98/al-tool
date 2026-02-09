#!/usr/bin/env python3
"""Streamed spreadsheet converter: .xlsb/.xlsx -> Apache Arrow IPC format.

Apache Arrow IPC format provides 10-100x faster read/write compared to JSONL:
- Binary columnar format optimized for analytics
- Zero-copy reads with memory-mapping
- Native type preservation (integers, floats, strings, dates)
- SIMD-friendly data layout

Output format: Arrow IPC Stream (.arrow file)
"""

from __future__ import annotations

import argparse
import logging
import sys
import gc
from decimal import Decimal
import datetime as _dt
from pathlib import Path
from typing import Any, List, Optional, Dict

import pyarrow as pa
from pyxlsb import open_workbook
from openpyxl import load_workbook


logger = logging.getLogger("xlsb_to_arrow")
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

# ============================================================================
# PERFORMANCE CONSTANTS - Optimized for 8GB RAM Windows machines
# ============================================================================
# Rows per batch for Arrow RecordBatch (larger = better compression, more memory)
BATCH_SIZE = 50000  # 50k rows per batch for optimal balance
# Progress logging interval
PROGRESS_LOG_INTERVAL = 50000  # Log every 50k rows
# Sample rows for schema inference
SCHEMA_SAMPLE_ROWS = 1000


def _normalize_value(v: Any) -> Any:
    """Normalize cell values to Python primitives.
    
    Arrow will handle type conversion automatically.
    We preserve original types as much as possible.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        # Keep as number - Arrow will use appropriate type
        return v
    if isinstance(v, Decimal):
        # Convert Decimal to float for Arrow
        return float(v)
    if isinstance(v, (_dt.datetime, _dt.date)):
        # Keep datetime objects - Arrow handles them natively
        return v
    # Everything else becomes string
    return str(v) if v is not None else None


def _ensure_parent_dir(path: Path) -> None:
    parent = path.parent
    if not parent.exists():
        parent.mkdir(parents=True, exist_ok=True)


def _infer_arrow_type(values: List[Any]) -> pa.DataType:
    """Infer Arrow type from sample values."""
    has_float = False
    has_int = False
    has_string = False
    has_datetime = False
    has_date = False
    
    for v in values:
        if v is None:
            continue
        if isinstance(v, bool):
            # bool is subclass of int, check first
            continue
        if isinstance(v, float):
            has_float = True
        elif isinstance(v, int):
            has_int = True
        elif isinstance(v, _dt.datetime):
            has_datetime = True
        elif isinstance(v, _dt.date):
            has_date = True
        elif isinstance(v, str):
            has_string = True
    
    # Priority: if any string, use string (safest for mixed data)
    if has_string:
        return pa.string()
    if has_datetime:
        return pa.timestamp('us')  # microseconds
    if has_date:
        return pa.date32()
    if has_float:
        return pa.float64()
    if has_int:
        return pa.float64()  # Use float64 for safety – avoids truncating decimals when sample rows are all-integer but later rows have floats
    
    # Default to string if no data
    return pa.string()


def _create_schema_from_samples(
    header: List[str], 
    sample_rows: List[List[Any]]
) -> pa.Schema:
    """Create Arrow schema from header and sample data."""
    # Transpose sample rows to get values per column
    num_cols = len(header)
    column_values: List[List[Any]] = [[] for _ in range(num_cols)]
    
    for row in sample_rows:
        for i in range(min(len(row), num_cols)):
            column_values[i].append(row[i])
    
    # Infer types
    fields = []
    for i, col_name in enumerate(header):
        # Sanitize column name
        safe_name = col_name if col_name else f"col_{i}"
        if isinstance(safe_name, str):
            safe_name = safe_name.strip()
            if not safe_name:
                safe_name = f"col_{i}"
        else:
            safe_name = f"col_{i}"
        
        col_type = _infer_arrow_type(column_values[i] if i < len(column_values) else [])
        fields.append(pa.field(safe_name, col_type))
    
    return pa.schema(fields)


def _convert_row_to_typed(row: List[Any], schema: pa.Schema) -> List[Any]:
    """Convert row values to match schema types."""
    result = []
    for i, field in enumerate(schema):
        v = row[i] if i < len(row) else None
        if v is None:
            result.append(None)
            continue
        
        # Type coercion based on field type
        if pa.types.is_string(field.type):
            result.append(str(v) if v is not None else None)
        elif pa.types.is_floating(field.type):
            try:
                result.append(float(v) if v is not None else None)
            except (ValueError, TypeError):
                result.append(None)
        elif pa.types.is_integer(field.type):
            try:
                result.append(int(v) if v is not None else None)
            except (ValueError, TypeError):
                result.append(None)
        elif pa.types.is_timestamp(field.type):
            if isinstance(v, _dt.datetime):
                result.append(v)
            elif isinstance(v, _dt.date):
                result.append(_dt.datetime.combine(v, _dt.time.min))
            else:
                result.append(None)
        elif pa.types.is_date(field.type):
            if isinstance(v, (_dt.datetime, _dt.date)):
                result.append(v if isinstance(v, _dt.date) else v.date())
            else:
                result.append(None)
        else:
            result.append(v)
    
    return result


def convert_xlsb_to_arrow(
    inpath: str,
    outpath: str,
    sheet_index: int = 1,
    header_row: int = 1,
    start_col: int = 1,
) -> Dict[str, Any]:
    """Convert XLSB to Arrow IPC format with streaming for memory efficiency.

    Args:
        header_row: 1-based row number where the header is located.
        start_col: 1-based column number from which to start reading.
    """
    in_path = Path(inpath)
    out_path = Path(outpath)
    _ensure_parent_dir(out_path)

    rows_written = 0
    header: List[str] = []
    sample_rows: List[List[Any]] = []
    schema: Optional[pa.Schema] = None
    writer: Optional[pa.ipc.RecordBatchStreamWriter] = None
    current_batch: List[List[Any]] = []
    # Convert 1-based start_col to 0-based index for slicing
    start_col_idx = max(0, start_col - 1)
    
    try:
        with open_workbook(str(in_path)) as wb:
            sheet_names = list(wb.sheets)
            if sheet_index < 1 or sheet_index > len(sheet_names):
                raise ValueError(f"sheet_index {sheet_index} out of range (1..{len(sheet_names)})")
            sheet_name = sheet_names[sheet_index - 1]
            logger.info(f"Converting sheet '{sheet_name}' from {in_path.name} to Arrow IPC")
            
            with wb.get_sheet(sheet_name) as sheet:
                current_row_num = 0  # 0-based counter
                header_found = False
                
                for row in sheet.rows():
                    current_row_num += 1  # now 1-based
                    
                    # Skip rows before the header row
                    if current_row_num < header_row:
                        continue
                    
                    row_values = []
                    for cell in row:
                        if cell is None or getattr(cell, "v", None) is None:
                            row_values.append(None)
                        else:
                            row_values.append(_normalize_value(cell.v))
                    
                    # Slice columns from start_col onwards
                    row_values = row_values[start_col_idx:]
                    
                    # Header row
                    if not header_found:
                        header = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(row_values)]
                        header_found = True
                        continue
                    
                    # Collect sample rows for schema inference
                    if len(sample_rows) < SCHEMA_SAMPLE_ROWS:
                        sample_rows.append(row_values)
                        continue
                    
                    # Create schema and writer after collecting samples
                    if schema is None:
                        schema = _create_schema_from_samples(header, sample_rows)
                        sink = pa.OSFile(str(out_path), 'wb')
                        writer = pa.ipc.new_stream(sink, schema)
                        
                        # Write sample rows as first batch
                        for sr in sample_rows:
                            current_batch.append(_convert_row_to_typed(sr, schema))
                            rows_written += 1
                        sample_rows.clear()
                    
                    # Add current row to batch
                    current_batch.append(_convert_row_to_typed(row_values, schema))
                    rows_written += 1
                    
                    # Write batch when full
                    if len(current_batch) >= BATCH_SIZE:
                        _write_batch(writer, schema, current_batch)
                        current_batch.clear()
                        
                        if rows_written % PROGRESS_LOG_INTERVAL == 0:
                            logger.info(f"Processed {rows_written:,} rows...")
                
                # Handle case where we never got enough samples
                if schema is None and (header or sample_rows):
                    if not header and sample_rows:
                        # First sample row becomes header
                        header = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(sample_rows[0])]
                        sample_rows = sample_rows[1:]
                    
                    schema = _create_schema_from_samples(header, sample_rows)
                    sink = pa.OSFile(str(out_path), 'wb')
                    writer = pa.ipc.new_stream(sink, schema)
                    
                    for sr in sample_rows:
                        current_batch.append(_convert_row_to_typed(sr, schema))
                        rows_written += 1
                
                # Write remaining batch
                if writer and current_batch:
                    _write_batch(writer, schema, current_batch)
                    current_batch.clear()
    
    finally:
        if writer:
            writer.close()
    
    # Force garbage collection after large conversion
    gc.collect()
    logger.info(f"Conversion complete: {rows_written:,} rows written to {out_path.name}")
    
    return {
        "status": "ok",
        "rows_written": rows_written,
        "output": str(out_path),
        "format": "arrow_ipc"
    }


def convert_xlsx_to_arrow(
    inpath: str,
    outpath: str,
    sheet_index: int = 1,
    header_row: int = 1,
    start_col: int = 1,
) -> Dict[str, Any]:
    """Convert XLSX to Arrow IPC format with optimized memory usage.

    Args:
        header_row: 1-based row number where the header is located.
        start_col: 1-based column number from which to start reading.
    """
    in_path = Path(inpath)
    out_path = Path(outpath)
    _ensure_parent_dir(out_path)

    rows_written = 0
    header: List[str] = []
    sample_rows: List[List[Any]] = []
    schema: Optional[pa.Schema] = None
    writer: Optional[pa.ipc.RecordBatchStreamWriter] = None
    current_batch: List[List[Any]] = []
    # Convert 1-based start_col to 0-based index for slicing
    start_col_idx = max(0, start_col - 1)
    
    try:
        # read_only=True and data_only=True are critical for memory efficiency
        wb = load_workbook(filename=str(in_path), read_only=True, data_only=True)
        sheets = wb.sheetnames
        if sheet_index < 1 or sheet_index > len(sheets):
            raise ValueError(f"sheet_index {sheet_index} out of range (1..{len(sheets)})")
        sheet = wb[sheets[sheet_index - 1]]
        logger.info(f"Converting sheet '{sheets[sheet_index - 1]}' from {in_path.name} to Arrow IPC")
        
        current_row_num = 0  # 0-based counter
        header_found = False
        
        for row in sheet.iter_rows(values_only=True):
            current_row_num += 1  # now 1-based
            
            # Skip rows before the header row
            if current_row_num < header_row:
                continue
            
            row_values = [_normalize_value(v) for v in row]
            
            # Slice columns from start_col onwards
            row_values = row_values[start_col_idx:]
            
            # Header row
            if not header_found:
                header = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(row_values)]
                header_found = True
                continue
            
            # Collect sample rows for schema inference
            if len(sample_rows) < SCHEMA_SAMPLE_ROWS:
                sample_rows.append(row_values)
                continue
            
            # Create schema and writer after collecting samples
            if schema is None:
                schema = _create_schema_from_samples(header, sample_rows)
                sink = pa.OSFile(str(out_path), 'wb')
                writer = pa.ipc.new_stream(sink, schema)
                
                # Write sample rows as first batch
                for sr in sample_rows:
                    current_batch.append(_convert_row_to_typed(sr, schema))
                    rows_written += 1
                sample_rows.clear()
            
            # Add current row to batch
            current_batch.append(_convert_row_to_typed(row_values, schema))
            rows_written += 1
            
            # Write batch when full
            if len(current_batch) >= BATCH_SIZE:
                _write_batch(writer, schema, current_batch)
                current_batch.clear()
                
                if rows_written % PROGRESS_LOG_INTERVAL == 0:
                    logger.info(f"Processed {rows_written:,} rows...")
        
        # Handle case where we never got enough samples
        if schema is None and (header or sample_rows):
            if not header and sample_rows:
                header = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(sample_rows[0])]
                sample_rows = sample_rows[1:]
            
            schema = _create_schema_from_samples(header, sample_rows)
            sink = pa.OSFile(str(out_path), 'wb')
            writer = pa.ipc.new_stream(sink, schema)
            
            for sr in sample_rows:
                current_batch.append(_convert_row_to_typed(sr, schema))
                rows_written += 1
        
        # Write remaining batch
        if writer and current_batch:
            _write_batch(writer, schema, current_batch)
            current_batch.clear()
        
        # Close workbook to release memory
        wb.close()
    
    finally:
        if writer:
            writer.close()
    
    # Force garbage collection after large conversion
    gc.collect()
    logger.info(f"Conversion complete: {rows_written:,} rows written to {out_path.name}")
    
    return {
        "status": "ok",
        "rows_written": rows_written,
        "output": str(out_path),
        "format": "arrow_ipc"
    }


def _write_batch(
    writer: pa.ipc.RecordBatchStreamWriter, 
    schema: pa.Schema, 
    rows: List[List[Any]]
) -> None:
    """Write a batch of rows to the Arrow stream."""
    if not rows:
        return
    
    # Transpose rows to columns
    num_cols = len(schema)
    columns = [[] for _ in range(num_cols)]
    
    for row in rows:
        for i in range(num_cols):
            columns[i].append(row[i] if i < len(row) else None)
    
    # Create Arrow arrays
    arrays = []
    for i, field in enumerate(schema):
        try:
            arr = pa.array(columns[i], type=field.type)
        except (pa.ArrowInvalid, pa.ArrowTypeError):
            # Fallback to string if type conversion fails
            arr = pa.array([str(v) if v is not None else None for v in columns[i]], type=pa.string())
        arrays.append(arr)
    
    # Create and write RecordBatch
    batch = pa.record_batch(arrays, schema=schema)
    writer.write_batch(batch)


def main(argv: list[str] | None = None) -> int:
    argv = list(argv) if argv is not None else sys.argv
    p = argparse.ArgumentParser(description="Convert .xlsb/.xlsx to Apache Arrow IPC format (streaming)")
    p.add_argument("infile", help="Input file (.xlsb or .xlsx)")
    p.add_argument("outfile", help="Output file (.arrow)")
    p.add_argument("--sheet", type=int, default=1, help="1-based sheet index to convert")
    p.add_argument("--header-row", type=int, default=1, help="1-based row number where the header is located (default: 1)")
    p.add_argument("--start-col", type=int, default=1, help="1-based column number from which to start reading (default: 1)")
    args = p.parse_args(argv[1:])

    infile_lower = args.infile.lower()
    conv_kwargs = dict(sheet_index=args.sheet, header_row=args.header_row, start_col=args.start_col)
    try:
        if infile_lower.endswith(".xlsb"):
            result = convert_xlsb_to_arrow(args.infile, args.outfile, **conv_kwargs)
        elif infile_lower.endswith(".xlsx") or infile_lower.endswith(".xls"):
            result = convert_xlsx_to_arrow(args.infile, args.outfile, **conv_kwargs)
        else:
            # Fallback: try xlsb reader first, else attempt openpyxl
            try:
                result = convert_xlsb_to_arrow(args.infile, args.outfile, **conv_kwargs)
            except Exception:
                result = convert_xlsx_to_arrow(args.infile, args.outfile, **conv_kwargs)
        
        # Output result as JSON for the worker to parse
        import json
        print(json.dumps(result))
        
    except Exception as exc:
        logger.exception("Conversion failed: %s", exc)
        import json
        print(json.dumps({"status": "error", "error": str(exc)}))
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
